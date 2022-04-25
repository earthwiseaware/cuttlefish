import unittest

from openpyxl import Workbook

from ..helpers import (
    get_columns,
    read_choices,
    write_choices
)

class TestWriteChoices(unittest.TestCase):
    def test_basic(self):
        choices = {
            'yes_no': {
                'yes': 'Yes',
                'no': 'No'
            },
            'degree': {
                'low': 'Low',
                'medium': 'Medium',
                'high': 'High'
            }
        }
        workbook = Workbook()
        sheet = workbook.active
        write_choices(sheet, choices)
        columns = {
            value: key for key, value in get_columns(sheet).items()
        }
        for row in sheet.iter_rows(min_row=2):
            key = row[columns['list_name']].value
            name = row[columns['name']].value
            assert row[columns['label']].value == choices[key][name]

class TestReadChoices(unittest.TestCase):
    def test_basic(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='list_name')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='yes_no')
        sheet.cell(row=2, column=2, value='yes')
        sheet.cell(row=2, column=3, value='Yes')
        # second data row
        sheet.cell(row=3, column=1, value='yes_no')
        sheet.cell(row=3, column=2, value='no')
        sheet.cell(row=3, column=3, value='No')
        # third data row
        sheet.cell(row=4, column=1, value='degree')
        sheet.cell(row=4, column=2, value='low')
        sheet.cell(row=4, column=3, value='Low')

        choices = read_choices(sheet)
        expected_choices = {
            'yes_no': {
                'yes': 'Yes',
                'no': 'No'
            },
            'degree': {
                'low': 'Low'
            }
        }
        assert choices == expected_choices

    def test_empty_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='list_name')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=3, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='yes_no')
        sheet.cell(row=2, column=2, value='yes')
        sheet.cell(row=2, column=3, value='Yes')
        # second data row
        sheet.cell(row=3, column=1, value='yes_no')
        sheet.cell(row=3, column=2, value='no')
        sheet.cell(row=3, column=3, value='No')
        # third data row
        sheet.cell(row=5, column=1, value='degree')
        sheet.cell(row=5, column=2, value='low')
        sheet.cell(row=5, column=3, value='Low')

        choices = read_choices(sheet)
        expected_choices = {
            'yes_no': {
                'yes': 'Yes',
                'no': 'No'
            },
            'degree': {
                'low': 'Low'
            }
        }
        assert choices == expected_choices

    def test_empty_columns(self):
        workbook = Workbook()
        sheet = workbook.active
        # setup columns
        sheet.cell(row=1, column=1, value='list_name')
        sheet.cell(row=1, column=2, value='name')
        sheet.cell(row=1, column=4, value='label')
        # first data row
        sheet.cell(row=2, column=1, value='yes_no')
        sheet.cell(row=2, column=2, value='yes')
        sheet.cell(row=2, column=4, value='Yes')
        # second data row
        sheet.cell(row=3, column=1, value='yes_no')
        sheet.cell(row=3, column=2, value='no')
        sheet.cell(row=3, column=4, value='No')
        # third data row
        sheet.cell(row=4, column=1, value='degree')
        sheet.cell(row=4, column=2, value='low')
        sheet.cell(row=4, column=4, value='Low')

        choices = read_choices(sheet)
        expected_choices = {
            'yes_no': {
                'yes': 'Yes',
                'no': 'No'
            },
            'degree': {
                'low': 'Low'
            }
        }
        assert choices == expected_choices