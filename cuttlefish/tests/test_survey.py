import unittest

from openpyxl import Workbook

from ..helpers import (
    write_survey,
    read_survey,
    get_columns
)

class TestWriteSurvey(unittest.TestCase):
    def test_basic_add(self):
        survey = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'select_multiple degree',
                'label': 'How crazy are your plans?',
                'name': 'insanity',
                'hint': 'Honesty is essential'
            }
        ]
        workbook = Workbook()
        sheet = workbook.active
        write_survey(sheet, survey)
        columns = get_columns(sheet)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            for j, cell in enumerate(row):
                assert cell.value == survey[i].get(columns[j], None)

    def test_single_recursion(self):
        survey = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group',
                'survey': [
                    {
                        'type': 'select_multiple degree',
                        'label': 'How crazy are your plans?',
                        'name': 'insanity',
                        'hint': 'Honesty is essential'
                    }
                ]
            }
        ]
        expected_rows = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group'
            },
            {
                'type': 'select_multiple degree',
                'label': 'How crazy are your plans?',
                'name': 'insanity',
                'hint': 'Honesty is essential'
            },
            {
                'type': 'end group'
            }
        ]
        workbook = Workbook()
        sheet = workbook.active
        write_survey(sheet, survey)
        columns = get_columns(sheet)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            for j, cell in enumerate(row):
                assert cell.value == expected_rows[i].get(columns[j], None)

    def test_multiple_recursion(self):
        survey = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group',
                'survey': [
                    {
                        'type': 'select_one yes_no',
                        'label': 'Are you a squirrel?',
                        'name': 'is_squirrel'
                    },
                    {
                        'type': 'begin repeat',
                        'survey': [
                            {
                                'type': 'select_multiple degree',
                                'label': 'How crazy are your plans?',
                                'name': 'insanity',
                                'hint': 'Honesty is essential'
                            }
                        ]
                    }
                    
                ]
            }
        ]
        expected_rows = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group'
            },
            {
                'type': 'select_one yes_no',
                'label': 'Are you a squirrel?',
                'name': 'is_squirrel'
            },
            {
                'type': 'begin repeat'
            },
            {
                'type': 'select_multiple degree',
                'label': 'How crazy are your plans?',
                'name': 'insanity',
                'hint': 'Honesty is essential'
            },
            {
                'type': 'end repeat'
            },
            {
                'type': 'end group'
            }
        ]
        workbook = Workbook()
        sheet = workbook.active
        write_survey(sheet, survey)
        columns = get_columns(sheet)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            for j, cell in enumerate(row):
                assert cell.value == expected_rows[i].get(columns[j], None)

    def test_subsequent_recursion(self):
        survey = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group',
                'survey': [
                    {
                        'type': 'select_one yes_no',
                        'label': 'Are you a squirrel?',
                        'name': 'is_squirrel'
                    } 
                ]
            },
            {
                'type': 'begin repeat',
                'survey': [
                    {
                        'type': 'select_multiple degree',
                        'label': 'How crazy are your plans?',
                        'name': 'insanity',
                        'hint': 'Honesty is essential'
                    }
                ]
            }
        ]
        expected_rows = [
            {
                'type': 'select_one yes_no',
                'label': 'Are you a lizard?',
                'name': 'is_lizard'
            },
            {
                'type': 'begin group',
                'label': 'A Group!',
                'name': 'group'
            },
            {
                'type': 'select_one yes_no',
                'label': 'Are you a squirrel?',
                'name': 'is_squirrel'
            },
            {
                'type': 'end group'
            },
            {
                'type': 'begin repeat'
            },
            {
                'type': 'select_multiple degree',
                'label': 'How crazy are your plans?',
                'name': 'insanity',
                'hint': 'Honesty is essential'
            },
            {
                'type': 'end repeat'
            }
        ]
        workbook = Workbook()
        sheet = workbook.active
        write_survey(sheet, survey)
        columns = get_columns(sheet)
        for i, row in enumerate(sheet.iter_rows(min_row=2)):
            for j, cell in enumerate(row):
                assert cell.value == expected_rows[i].get(columns[j], None)
